#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_context.py — Phase 0 preparation for the IM-110 spec auto-generation.

Does three things, all read-only against the source materials:
  1. Extract screen mockup images from docs/specs/IM-110操作表示一覧_*.xlsx
     (the 400x240 monochrome LCD screens), save PNG + BMP into
     docs/specs/assets/im-110/, and emit an image manifest JSON.
  2. Extract the ID-200T spec text (structural template) via `textutil`.
  3. Assemble a plain-text "fact pack" (CLAUDE.md §4, protocol summary,
     xlsx sheet labels) so the generation step feeds the LLM facts only.

Outputs land under tools/spec-gen/context/ and docs/specs/assets/im-110/.

Run:  python3 tools/spec-gen/build_context.py
"""
import json
import os
import subprocess
import sys
from pathlib import Path

import openpyxl
from PIL import Image, ImageOps

# thin black frame (px) drawn around each 400x240 screen so the LCD extent is
# visible when the spec is printed (white LCD content otherwise blends into paper)
BORDER_PX = 2

REPO = Path(__file__).resolve().parents[2]          # IM-110-system/
SPECS = REPO / "docs" / "specs"
ASSETS = SPECS / "assets" / "im-110"
CTX = REPO / "tools" / "spec-gen" / "context"

XLSX = SPECS / "IM-110操作表示一覧_20260507.xlsx"
DOCX = SPECS / "ID-200T プログラム仕様書V13.docx"

# sheet title (as stored in the workbook) -> short slug used for filenames
SHEET_SLUG = {
    "1．測定": "meas",
    "2．校正": "calib",
    "3．設定": "settings",
}
# Only images at/above this size are treated as real LCD screens (400x240).
# Smaller anchored images are flow-diagram icons/arrows and are skipped.
SCREEN_MIN_W = 200
SCREEN_MIN_H = 120


def extract_screen_images() -> list[dict]:
    """Extract 400x240 screen mockups from the xlsx, save PNG+BMP, return manifest."""
    ASSETS.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(XLSX)
    manifest: list[dict] = []
    for ws in wb.worksheets:
        slug = SHEET_SLUG.get(ws.title)
        if slug is None:
            continue
        imgs = getattr(ws, "_images", [])
        # order by anchor (row, then col) to approximate the flow reading order
        def anchor_rc(img):
            frm = getattr(img.anchor, "_from", None)
            return (getattr(frm, "row", 0), getattr(frm, "col", 0))

        screens = [im for im in imgs
                   if (im.width or 0) >= SCREEN_MIN_W and (im.height or 0) >= SCREEN_MIN_H]
        screens.sort(key=anchor_rc)
        for idx, img in enumerate(screens, 1):
            frm = getattr(img.anchor, "_from", None)
            row = getattr(frm, "row", None)
            col = getattr(frm, "col", None)
            data = img._data()
            base = f"{slug}_{idx:02d}"
            png_path = ASSETS / f"{base}.png"
            bmp_path = ASSETS / f"{base}.bmp"
            png_path.write_bytes(data)
            # add a thin black frame so the 400x240 LCD extent is visible in print,
            # then write both PNG and BMP framed (same filenames the spec links).
            with Image.open(png_path) as im:
                framed = ImageOps.expand(im.convert("RGB"), border=BORDER_PX, fill=(0, 0, 0))
                framed.save(png_path, format="PNG")
                framed.save(bmp_path, format="BMP")
            manifest.append({
                "id": base,
                "sheet": ws.title,
                "slug": slug,
                "order": idx,
                "anchor_row": row,
                "anchor_col": col,
                "w": img.width,
                "h": img.height,
                "png": str(png_path.relative_to(SPECS)),
                "bmp": str(bmp_path.relative_to(SPECS)),
            })
    return manifest


def extract_id200t_text() -> str:
    """Convert the ID-200T docx to plain text as the structural template."""
    out = CTX / "id200t_template.txt"
    subprocess.run(
        ["textutil", "-convert", "txt", "-output", str(out), str(DOCX)],
        check=True,
    )
    return out.read_text(encoding="utf-8", errors="ignore")


def dump_xlsx_labels() -> dict:
    """Pull the sparse cell labels from each sheet (mode / menu structure)."""
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    out = {}
    for ws in wb.worksheets:
        if ws.title not in SHEET_SLUG:
            continue
        labels = []
        for r in ws.iter_rows():
            for c in r:
                v = c.value
                if v not in (None, "") and str(v).strip():
                    labels.append({"cell": c.coordinate, "text": str(v).strip()})
        out[ws.title] = labels
    return out


def main() -> int:
    CTX.mkdir(parents=True, exist_ok=True)
    print("[1/3] extracting screen images from xlsx ...")
    manifest = extract_screen_images()
    (CTX / "image_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    by_slug = {}
    for m in manifest:
        by_slug.setdefault(m["slug"], 0)
        by_slug[m["slug"]] += 1
    print(f"      {len(manifest)} screens -> {ASSETS}  ({by_slug})")

    print("[2/3] extracting ID-200T template text ...")
    tpl = extract_id200t_text()
    print(f"      {len(tpl.splitlines())} lines -> context/id200t_template.txt")

    print("[3/3] dumping xlsx labels ...")
    labels = dump_xlsx_labels()
    (CTX / "xlsx_labels.json").write_text(
        json.dumps(labels, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"      {sum(len(v) for v in labels.values())} labels -> context/xlsx_labels.json")

    print("done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
